"""

"""
import codecs
import csv
import os
import sys
from openpyxl import load_workbook


def trimrow(r: list, maxlen: int):
    """

    :param r:
    :param maxlen:
    :return: Trailing blank cells are removed from the list
    """
    ix = len(r) - 1
    while len(r) > maxlen:
        if r[ix]:
            break
        r.pop()
        ix -= 1


def row_dict_reader(filename: str | None, verbos=1, skiprows=0,
                    allow_long_rows=False):
    """
    A generator function to iterate through either a CSV file or XLSX file.

    We only look at the first sheet of an XLSX file.

    :param filename:
    :param verbos:
    :param skiprows:
    :param allow_long_rows:
    :return: an iterator that calls this function or None
    """
    if not filename:
        return None
    _, suffix = os.path.splitext(filename)
    if suffix.lower() == '.csv':
        with codecs.open(filename, 'r', 'utf-8-sig') as mapfile:
            for _ in range(skiprows):
                next(mapfile)
            reader = csv.DictReader(mapfile)
            n_input_fields = len(reader.fieldnames)
            if verbos >= 1:
                print(f'CSV Column Headings: {", ".join(reader.fieldnames)}')
            for nrow, row in enumerate(reader):
                if not allow_long_rows and len(row) > n_input_fields:
                    print(f"Error: row {nrow + 1} longer than heading: {row}")
                    sys.exit(1)
                yield row
    elif suffix.lower() == '.xlsx':
        wb = load_workbook(filename=filename)
        ws = wb.active
        enumrows = enumerate(ws.iter_rows(values_only=True))
        for _ in range(skiprows):
            next(enumrows)
        _, heading = next(enumrows)
        heading = list(heading)  # tuple -> list so it can be trimmed
        trimrow(heading, 0)
        # print(len(heading))
        # sys.exit()
        n_input_fields = len(heading)
        if verbos >= 1:
            print(f'Excel Column Headings: '
                  f'{", ".join([str(x) for x in heading])}')
        for nrow, rawrow in enumrows:
            rawrow = list(rawrow)
            row = dict()
            # print(f'before: {rawrow}')
            trimrow(rawrow, n_input_fields)
            # print(f'after : {rawrow}')
            if not allow_long_rows and len(rawrow) > n_input_fields:
                print(f"Error: row {nrow + 1} longer than heading: {rawrow}")
                sys.exit(1)
            for ncell, cell in enumerate(rawrow):
                if type(cell) == str:
                    cell = cell.replace('\n', ' ')
                row[heading[ncell]] = '' if cell is None else str(cell).strip()
            if not ''.join(row.values()):
                continue
            yield row
