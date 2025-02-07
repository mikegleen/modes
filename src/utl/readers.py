"""

"""
import codecs
import csv
import datetime
import os
import sys
from openpyxl import load_workbook
# noinspection PyPep8Naming
import xml.etree.ElementTree as ET

from utl.cfgutil import Config
from utl.normalize import normalize_id
from utl.zipmagic import openfile


def _trimrow(r: list, maxlen: int):
    """

    :param r:
    :param maxlen:
    :return: Trailing blank cells are removed from the list
    """
    ix = len(r) - 1
    while len(r) > maxlen:
        if r[ix]:
            break
        r.pop()
        ix -= 1
    for n, elt in enumerate(r):
        if elt is None:
            r[n] = ''
        else:
            r[n] = ' '.join(str(elt).split())
    return r


def object_reader(infilename: str | None, config=None, normalize=False, verbos=1):
    """
    Open an XML file and return Object elements.

    :param infilename: The XML file in Modes format
    :param config: A Config object or None in which case an empty one is made.
                   The function refers to config members ``record_tag`` and ``record_id_xpath``.
    :param normalize: return a modified tuple (see below)
    :param verbos: trace level
    :return: An iterator that returns tuples of:
                (accession number, element)
             If normalize==True, return tuples of:
                (accession number, normalized accession number, element)
    """
    objectlevel = 0
    if config is None:
        config = Config()
    with openfile(infilename) as infile:
        for event, elem in ET.iterparse(infile, events=('start', 'end')):
            # print(event)
            if event == 'start':
                # print(elem.tag)
                if elem.tag == config.record_tag:
                    objectlevel += 1
                continue
            # It's an "end" event.
            if elem.tag != config.record_tag:  # default: Object
                continue
            objectlevel -= 1
            if objectlevel:
                continue  # It's not a top level Object.
            idelem = elem.find(config.record_id_xpath)
            idnum = idelem.text if idelem is not None else ''
            if verbos >= 3:
                print('idnum: {}', idnum)
            if normalize:
                yield idnum, normalize_id(idnum), elem
            else:
                yield idnum, elem


def get_heading(filepath: str | None, verbos=1, skiprows=0) -> list | None:
    """

    :param filepath: the path to the CSV or XLSX input file or None. If None,
                     None is returned.
    :param verbos: level of tracing. If > 1, print column headings.
    :param skiprows: Rows to skip at the front of the file. Default = 0.
    :return: a list of the heading row or None
    """
    if not filepath:
        return None
    _, suffix = os.path.splitext(filepath)
    if suffix.lower() == '.csv':
        with codecs.open(filepath, encoding='utf-8-sig') as mapfile:
            for _ in range(skiprows):
                next(mapfile)
            reader = csv.DictReader(mapfile)
            return list(reader.fieldnames)
    elif suffix.lower() == '.xlsx':
        wb = load_workbook(filename=filepath)
        ws = wb.active
        enumrows = enumerate(ws.iter_rows(values_only=True))
        for _ in range(skiprows):
            next(enumrows)
        _, heading = next(enumrows)
        heading = list(heading)  # tuple -> list so it can be trimmed
        _trimrow(heading, 0)
        n_input_fields = len(heading)
        # sys.exit()
        if verbos >= 1:
            print(f'------- get_heading: {filepath}, heading length: {n_input_fields}')
            print(f'Excel Column Headings: '
                  f'{", ".join([str(x) for x in heading])}')
        return heading
    else:
        if verbos >= 1:
            print(f'"{suffix}" is not a valid suffix. Terminating.')
        sys.exit(1)


def cleancell(cell) -> str:
    if type(cell) is str:
        cell = cell.replace('\n', ' ')
    elif cell is None:
        cell = ''
    elif isinstance(cell, datetime.datetime):
        cell = cell.strftime('%Y-%m-%d')
    else:
        cell = str(cell)
    return cell


def row_dict_reader(filename: str | None, verbos=1, skiprows=0,
                    allow_long_rows=False):
    """
    A generator function to iterate through either a CSV file or XLSX file.

    The input file must contain a heading row.
    We only look at the first sheet of an XLSX file.

    :param filename: The input CSV or XLSX file.
    :param verbos: Verbosity. If > 1, print column headings
    :param skiprows: Rows to skip at the front of the file. Default = 0.
    :param allow_long_rows: Normally, if a data row is longer than the heading
                            row, the program aborts. If this parameter is set
                            to True, this check is disabled.
    :return: an iterator that calls this function or None. Each iteration returns
             a dict where the keys are the column titles and the values are the
             column contents for each row.
    """
    if not filename:
        return None
    _, suffix = os.path.splitext(filename)
    if suffix.lower() == '.csv':
        with codecs.open(filename, encoding='utf-8-sig') as mapfile:
            for _ in range(skiprows):
                next(mapfile)
            reader = csv.DictReader(mapfile)
            n_input_fields = len(reader.fieldnames)
            if verbos >= 2:
                fieldnames = ", ".join(_trimrow(list(reader.fieldnames), 1))
                print(f'CSV Column Headings: {fieldnames}')
            for nrow, row in enumerate(reader, start=1):
                if not allow_long_rows and len(row) > n_input_fields:
                    print(f"Error: row {nrow} longer than heading: {row}")
                    sys.exit(1)
                yield row
    elif suffix.lower() == '.xlsx':
        wb = load_workbook(filename=filename)
        ws = wb.active
        enumrows = enumerate(ws.iter_rows(values_only=True))
        for _ in range(skiprows):
            next(enumrows)
        _, heading = next(enumrows)
        heading = list(heading)  # tuple -> list so it can be trimmed
        _trimrow(heading, 0)
        n_input_fields = len(heading)
        # sys.exit()
        if verbos >= 2:
            print(f'row_dict_reader heading length: {n_input_fields}')
            print(f'Excel Column Headings: '
                  f'{", ".join([str(x) for x in heading])}')
        for nrow, rawrow in enumrows:
            rawrow = list(rawrow)
            # print(f'{rawrow=}')
            row = dict()
            # print(f'before: {rawrow}')
            _trimrow(rawrow, n_input_fields)
            # print(f'after : {rawrow}')
            if not allow_long_rows and len(rawrow) > n_input_fields:
                print(f"Error: row {nrow + 1} longer than heading: {rawrow}")
                sys.exit(1)
            for ncell, cell in enumerate(rawrow):
                ccell = cleancell(cell)
                # print(f'{cell=}, {ccell=}')
                row[heading[ncell]] = ccell
                # if row['Accession Number'] == 'JB008':
                #     sys.exit(1)
            if not ''.join(row.values()):
                continue
            # print(f'{row=}')

            yield row
    else:
        if verbos >= 1:
            print(f'"{suffix}" is not a valid suffix. Terminating.')
        sys.exit(1)


def row_list_reader(filename: str | None, verbos=1, skiprows=0,
                    allow_long_rows=False):
    """
    A generator function to iterate through either a CSV file or XLSX file.
    We only look at the first sheet of an XLSX file.
    The first row is taken as the header.

    :param filename: The input CSV or XLSX file.
    :param verbos: Verbosity. If > 1, print column headings
    :param skiprows: Rows to skip at the front of the file. Default = 0.
    :param allow_long_rows: Normally, if a data row is longer than the heading
                            row, the program aborts. If this parameter is set
                            to True, this check is disabled.
    :return: an iterator that calls this function or None. Each iteration returns
             a row as a list.
    """
    if not filename:
        return None
    _, suffix = os.path.splitext(filename)
    if suffix.lower() == '.csv':
        with codecs.open(filename, encoding='utf-8-sig') as mapfile:
            reader = csv.reader(mapfile)
            for _ in range(skiprows):
                next(reader)
            header = next(reader)
            n_input_fields = len(header)
            if verbos >= 2:
                fieldnames = ", ".join(_trimrow(header, 1))
                print(f'CSV Column Headings: {fieldnames}')
            for nrow, row in enumerate(reader):
                if not allow_long_rows and len(row) > n_input_fields:
                    print(f"Error: row {nrow + 1} longer than heading: {row}")
                    sys.exit(1)
                yield row
    elif suffix.lower() == '.xlsx':
        wb = load_workbook(filename=filename)
        ws = wb.active
        enumrows = enumerate(ws.iter_rows(values_only=True))
        for _ in range(skiprows):
            next(enumrows)
        _, heading = next(enumrows)
        heading = list(heading)  # tuple -> list so it can be trimmed
        _trimrow(heading, 0)
        n_input_fields = len(heading)
        # sys.exit()
        if verbos >= 2:
            print(f'row_list_reader heading length: {n_input_fields}')
            print(f'Excel Column Headings: '
                  f'{", ".join([str(x) for x in heading])}')
        for nrow, rawrow in enumrows:
            rawrow = list(rawrow)
            _trimrow(rawrow, n_input_fields)
            rawrow = [cleancell(c) for c in rawrow]
            if not allow_long_rows and len(rawrow) > n_input_fields:
                print(f"Error: row {nrow + 1} longer than heading: {rawrow}")
                sys.exit(1)
            if not ''.join(rawrow):
                continue
            yield rawrow
    else:
        if verbos >= 1:
            print(f'"{suffix}" is not a valid suffix. Terminating.')
        sys.exit(1)
