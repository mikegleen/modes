"""
    Create a new spreadsheet with the columns defined by NEWCOLS. If any of the
    columns exist in the old spreadsheet copy those.

    If the names don’t match exactly, map the new to the old name in
    NEW_TO_OLD_MAP.

    If the output file is .xlsx, all cells are set to type string.

"""
import argparse
import codecs
import csv
import os.path
import sys

import openpyxl.cell.cell
from openpyxl import Workbook

from utl.normalize import sphinxify
from utl.readers import row_dict_reader

NEWCOLS = ('Serial,Pages,Date,Person From,Person To,Org From,Org To,Type,'
           + 'Comment,Description').split(',')

NEW_TO_OLD_MAP = {'Pages': 'Multiple Images',
                  'Person From': 'From',
                  'Person To': 'To'}


def onefile(filename):
    is_xlsx = filename.lower().endswith('.xlsx')
    outpath = os.path.join(_args.outdir, filename)
    workbook = ws = outcsv = None  # Stop pycharm whining
    if is_xlsx:
        workbook = Workbook()
        del workbook[workbook.sheetnames[0]]  # remove the default sheet
        ws = workbook.create_sheet('Sheet1')
        for col_num, col in enumerate(NEWCOLS, start=1):
            cell = ws.cell(row=1, column=col_num, value=col)
            cell.data_type = openpyxl.cell.cell.TYPE_STRING
    else:
        encoding = 'utf-8-sig'  # if _args.bom else 'utf-8'
        csvfile = codecs.open(outpath, 'w', encoding)
        outcsv = csv.writer(csvfile, delimiter=',')
        outcsv.writerow(NEWCOLS)
    infilepath = os.path.join(_args.indir, filename)
    row_num = 1
    for inrow in row_dict_reader(infilepath, _args.verbose,
                                 _args.skip_rows):
        row_num += 1
        outrow = []
        for col_num, col in enumerate(NEWCOLS, start=1):
            if col in inrow:
                value = inrow[col]
            elif col in NEW_TO_OLD_MAP:
                value = inrow[NEW_TO_OLD_MAP[col]]
            else:
                value = ''
            if value is None:
                value = ''
            outrow.append(value)
        if not ''.join(outrow[1:]):  # empty row
            continue
        # print(outrow)
        if is_xlsx:
            for col_num, value in enumerate(outrow, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # cell.data_type = openpyxl.cell.cell.TYPE_STRING
        else:
            outcsv.writerow(outrow)

    if is_xlsx:
        workbook.save(outpath)
    return


def main():
    os.makedirs(_args.outdir, exist_ok=True)
    infiles = os.listdir(_args.indir)
    for infile in infiles:
        _, suffix = os.path.splitext(infile)
        if suffix.lower() not in ('.csv', '.xlsx') or infile.startswith('~'):
            print(f'Skipping {infile}')
            continue
        if _args.verbose >= 2:
            print('***', infile)
        onefile(infile)
        if _args.short:
            return


def getparser():
    parser = argparse.ArgumentParser(description=sphinxify('''

                ''', calledfromsphinx))
    parser.add_argument('indir', help='''
        Input folder with CSV or XLSX files.
        ''')
    parser.add_argument('outdir', help='''
        Output folder to contain newly created files.
        ''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    parser.add_argument('-s', '--short', action='store_true', help='''
        Only process one object. For debugging.''')
    parser.add_argument('--skip_rows', type=int, default=0, help='''
        Skip rows at the beginning of the input file. In the output file,
        the heading will start at the first row.''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


calledfromsphinx = True


if __name__ == '__main__':
    calledfromsphinx = False
    assert sys.version_info >= (3, 9)
    if len(sys.argv) == 1:
        sys.argv.append('-h')
    _args = getargs(sys.argv)
    main()
