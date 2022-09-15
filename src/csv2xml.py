# -*- coding: utf-8 -*-
"""

"""
import argparse
import codecs
import copy
import csv
import os.path
import re
import sys
# noinspection PyPep8Naming
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


from utl.cfgutil import Config, Stmt, Cmd, new_subelt
from utl.normalize import modesdatefrombritishdate, sphinxify, if_not_sphinx
from utl.normalize import DEFAULT_MDA_CODE, normalize_id, denormalize_id


def trace(level, template, *args):
    if _args.verbose >= level:
        print(template.format(*args))


def clean_accnum(accnum: str):
    """

    :param accnum: input alleged accession number. Remove any leading zeroes
                   and convert back to denormalized form fit for insertion into
                   the XML file
    :return:       denormalized accession number
                   Raises ValueError or AssertionError if the accession number
                   is invalid.
    """
    naccnum = normalize_id(accnum)
    return denormalize_id(naccnum)


def row_reader(verbos, skiprows):
    """
    A generator function to iterate throw either a CSV file or XLSX file.

    :return:
    """
    filename = _args.incsvfile
    _, suffix = os.path.splitext(filename)
    if suffix.lower() == '.csv':
        with codecs.open(filename, 'r', 'utf-8-sig') as mapfile:
            for _ in range(skiprows):
                next(mapfile)
            reader = csv.DictReader(mapfile)
            if verbos >= 1:
                print(f'CSV Column Headings: {reader.fieldnames}')
            for row in reader:
                yield row
    elif suffix.lower() == '.xlsx':
        wb = load_workbook(filename=filename)
        ws = wb.active
        enumrows = enumerate(ws.iter_rows(values_only=True))
        for _ in range(skiprows):
            next(enumrows)
        _, heading = next(enumrows)
        if verbos >= 1:
            print(f'Excel Column Headings: {heading}')
        for nrow, rawrow in enumrows:
            row = dict()
            for ncell, cell in enumerate(rawrow):
                if type(cell) == str:
                    cell = cell.replace('\n', ' ')
                row[heading[ncell]] = cell
            yield row


def next_accnum(accnum: str):
    """
    :param accnum: An accession number like LDHRM.2021.2
    :return: The accession number with the trailing index incremented after
             each call.
    """
    if not accnum:
        return None
    mat = re.match(r'(.+?)(\d+$)', accnum)
    prefix = mat.group(1)
    suffix = int(mat.group(2))
    while True:
        yield prefix + str(suffix)
        suffix += 1


def create_items(doc, elt, root, text):
    """
    :param doc: the Items document from the config
    :param elt: the element created from the xpath
    :param root: the object element
    :param text: The descriptions joined by the delimiter usually "|"
    :return: None
    """
    olditems = elt.findall('Item')
    for item in olditems:
        elt.remove(item)
    if not text:
        return
    itemtexts = text.split(doc[Stmt.MULTIPLE_DELIMITER])
    itemnumber = 0
    for itemnumber, text in enumerate(itemtexts, start=1):
        item = ET.SubElement(elt, 'Item')
        listnumber = ET.SubElement(item, 'ListNumber')
        listnumber.text = str(itemnumber)
        description = ET.SubElement(item, 'BriefDescription')
        description.text = text
    numberofitems = root.find('NumberOfItems')
    if numberofitems is None:
        numberofitems = ET.SubElement(root, "NumberOfItems")
    numberofitems.text = str(itemnumber)


def get_object_from_file(templatefilepath):
    """
    :param templatefilepath: The path from the --template parameter or the csv
                             file
    :return: the Object element
    """
    with open(templatefilepath) as templatefile:
        root = ET.parse(templatefile)
        # Assume that the template is actually an Interchange file with empty
        # elements.
        object_template = root.find(config.record_tag)
        if object_template is None:
            # Ok, so maybe it really is a template
            object_template = root.find(f'./template/{config.record_tag}')
            if object_template is None:
                raise ValueError(f'Cannot find the {config.record_tag} element'
                                 f' from root or from ./template')
    return object_template


def get_template_from_csv(row: dict[str]):
    key = row[config.template_title]
    if key not in config.templates:
        raise ValueError(f'Template key in CSV file: {key} is not in config. {row=}')
    templatefilepath = os.path.join(config.template_dir, config.templates[key])
    trace(2, 'template file: {}', templatefilepath)
    return get_object_from_file(templatefilepath)


def main():
    global nrows
    if not _args.noprolog:
        outfile.write(b'<?xml version="1.0"?><Interchange>\n')
    global_object_template = None
    if _args.template:
        global_object_template = get_object_from_file(_args.template)
    reader = row_reader(_args.verbos, _args.skip_rows)
    nrows = 0
    # PyCharm whines if we don't initialize accnumgen
    accnumgen = next_accnum(_args.acc_num)
    for row in reader:
        emit = True
        if global_object_template:
            template = copy.deepcopy(global_object_template)
        else:
            template = get_template_from_csv(row)
        elt = template.find(config.record_id_xpath)
        if _args.acc_num:
            accnum = next(accnumgen)
            trace(2, 'Serial generated: {}', accnum)
        else:
            accnum = row[_args.serial]
            if config.add_mda_code and accnum[0].isnumeric():
                accnum = _args.mdacode + '.' + accnum
        accnum = clean_accnum(accnum)
        elt.text = accnum
        for doc in config.col_docs:
            cmd = doc[Stmt.CMD]
            title = doc[Stmt.TITLE]
            if cmd != Cmd.CONSTANT and not row[title]:
                trace(3, '{}: cell empty {}', accnum, title)
                continue
            xpath = doc[Stmt.XPATH]
            elt = template.find(xpath)
            if elt is None:
                elt = new_subelt(doc, template, accnum, _args.verbose)
            if elt is None:
                trace(1, '{}: Cannot create new {}.\nCheck parent_path statement.',
                      accnum, doc[Stmt.XPATH])
                continue
            if cmd == Cmd.CONSTANT:
                elt.text = doc[Stmt.VALUE]
                continue
            text = row[title]
            if Stmt.REQUIRED in doc and not text:
                print(f'*** Required column “{title}” is missing from'
                      f' {accnum}. Object excluded.')
                emit = False
            if cmd == Cmd.ITEMS:
                create_items(doc, elt, template, text)
            else:
                elt.text = text
            if Stmt.DATE in doc:
                # Only britishdate supported now
                try:
                    elt.text, _, _ = modesdatefrombritishdate(elt.text)
                    # print(type(elt.text))
                except ValueError:
                    elt.text = 'unknown'
        if emit:
            nrows += 1
            outfile.write(ET.tostring(template))
    if not _args.noprolog:
        outfile.write(b'</Interchange>')


def getparser():
    parser = argparse.ArgumentParser(description=sphinxify('''
        Read a CSV file containing two or more columns. The first column
        is the accession number and the following columns are the fields
        defined by the XPATH statement in the config file. The first row in the
        CSV file is a heading row. The column titles must match the document
        titles in the config file. Columns are referred to by name so it is
        permissible to omit columns from the config file.
        
        Create an
        XML file with data from the CSV file based on a template of the
        XML structure.
                ''', calledfromsphinx))
    parser.add_argument('--acc_num', help='''
        This is the first accession number in a series to assign to rows
        in the input CSV file. If specified, the column in the CSV file
        containing an accession number, if one exists, is ignored. For example,
        if the parameter is "LDHRM.2021.2", the numbers assigned will be
        "LDHRM.2021.2", "LDHRM.2021.3", etc. This value will be stored in the
        ObjectIdentity/Number element.''')
    parser.add_argument('-c', '--cfgfile', required=True,
                        type=argparse.FileType('r'), help=sphinxify('''
        The YAML file describing the column path(s) to update.
        The config file may contain only ``column``, ``constant``, or
        ``items`` commands.
    ''', calledfromsphinx))
    parser.add_argument('-i', '--incsvfile', help=sphinxify('''
        The file containing data to be inserted into the XML template.
        The file must have a heading, but see option --skip_rows.
        The heading of the column containing the serial number must be
        ``Serial`` or an alternative set by --serial parameter. Subsequent
        columns must match the corresponding title in the configuration file.
        The file must have a (case-insensitive) suffix of ``.csv`` or ``.xlsx``
        containing CSV format or Excel format data, respectively.''',
                                                            calledfromsphinx))
    parser.add_argument('-m', '--mdacode', default=DEFAULT_MDA_CODE, help=f'''
        Specify the MDA code, used in normalizing the accession number.''' +
                        if_not_sphinx(''' The default is "{DEFAULT_MDA_CODE}".
                        ''', calledfromsphinx))
    parser.add_argument('-o', '--outfile', help='''
        The output XML file.''')
    parser.add_argument('-p', '--noprolog', action='store_true', help='''
        Inhibit the insertion of an XML prolog at the front of the file and an
        <Interchange> element as the root. This results in an invalid XML file
        but is useful if the output is to be manually edited.''')
    parser.add_argument('--serial', default='Serial', help=sphinxify('''
        The column containing the serial number (the first column) must have a
        heading with this value. This is ignored if the --acc_num parameter is
        specified.
        ''' + if_not_sphinx('''
        The default value is "Serial".''', calledfromsphinx),
                                       calledfromsphinx))
    parser.add_argument('-s', '--short', action='store_true', help='''
        Only process one object. For debugging.''')
    parser.add_argument('--skip_rows', type=int, default=0, help='''
        Skip rows at the beginning of the CSV file.''')
    parser.add_argument('-t', '--template', help=sphinxify('''
        The XML file that is the template for creating the output XML.
        Specify this or global statements in the configuration
        ``template_dir``, ``template_title``, and ``templates``.
        ''', calledfromsphinx))
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


def check_cfg(c):
    """
    Only the column command is allowed.
    :param c: The Config instance
    :return: Nonzero if the config fails.
    """
    errs = 0
    for doc in c.col_docs:
        if doc[Stmt.CMD] not in (Cmd.COLUMN, Cmd.CONSTANT, Cmd.ITEMS):
            print(f'Command "{doc[Stmt.CMD]}" not allowed, exiting')
            errs += 1
    for doc in c.ctrl_docs:
        print(f'Command "{doc[Stmt.CMD]}" not allowed, exiting.')
        errs += 1
    return errs


calledfromsphinx = True


if __name__ == '__main__':
    global nrows
    assert sys.version_info >= (3, 9)
    calledfromsphinx = False
    _args = getargs(sys.argv)
    incsvfile = codecs.open(_args.incsvfile, 'r', encoding='utf-8-sig')
    outfile = open(_args.outfile, 'wb')
    trace(1, 'Input file: {}', _args.incsvfile)
    trace(1, 'Creating file: {}', _args.outfile)
    config: Config = Config(_args.cfgfile, dump=_args.verbose > 1,
                            allow_required=True)
    if errors := check_cfg(config):
        trace(1, '{} errors found. Aborting.', errors)
        sys.exit(1)
    main()
    trace(1, 'End csv2xml. {} object{} written.', nrows,
          '' if nrows == 1 else 's')
