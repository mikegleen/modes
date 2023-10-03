# -*- coding: utf-8 -*-
"""
    Create a report of box contents with columns for the stocktake.
    Parameters:
        1. Input XML file. This is a Modes database.
        2. Output CSV or XLSX file depending on the file extension of the output
           file.
"""
import argparse
import codecs
from collections import defaultdict
import csv
import io
import re
import sys

import openpyxl.cell.cell
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook

from utl.cfgutil import Config, Stmt, Cmd
from utl.normalize import sphinxify, normalize_id
from utl.readers import object_reader

CFG_STRING = """
cmd: column
xpath: ./ObjectLocation[@elementtype="normal location"]/Location
title: Normal
---
cmd: column
xpath: ./Identification/Title
width: 50
---
cmd: constant
xpath: dummy
title: Condition
value:
"""


def pad_loc(loc):
    """
    If the location field matches the pattern "<Letter(s)><Number(s)>" like
    "S1" then pad the number part so that it sorts correctly. Also convert the
    letter part to upper case
    :param loc: location
    :return: padded location
    """
    m = re.match(r'(\D+)(\d+)', loc)
    if m:
        g1 = m.group(1).upper()
        g2 = int(m.group(2))
        return f'{g1}{g2:03}'
    else:
        return loc


def unpad_loc(loc):
    m = re.match(r'(\D+)(\d+)', loc)
    if m:
        g1 = m.group(1).upper()
        g2 = int(m.group(2))
        return f'{g1}{g2}'
    else:
        return loc


def one_xml_object(elt):
    num = elt.find('./ObjectIdentity/Number').text
    loc = elt.find('./ObjectLocation[@elementtype="current location"]/Location')
    if loc is not None and loc.text:
        location = pad_loc(loc.text)
    else:
        location = 'unknown'
    row = [num]
    for doc in cfg.col_docs:
        if doc[Stmt.CMD] != Cmd.COLUMN:
            continue
        xpath = doc[Stmt.XPATH]
        data = elt.find(xpath)
        if data is not None and data.text:
            datatext = data.text
        else:
            datatext = 'unknown'
        if Stmt.WIDTH in doc:
            datatext = datatext[:int(doc[Stmt.WIDTH])]
        row.append(datatext)
    boxdict[location].append(row)


def main(config):
    for _, elt in object_reader(_args.infile, config=config):
        one_xml_object(elt)
        elt.clear()
    nrow = -1
    for box in sorted(boxdict.keys()):
        if is_xlsx:
            nrow += 2
            cell = ws.cell(row=nrow, column=1, value=unpad_loc(box))
            cell.data_type = openpyxl.cell.cell.TYPE_STRING
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            ncol = 1
            for doc in cfg.col_docs:
                ncol += 1
                cell = ws.cell(row=nrow, column=ncol, value=doc['title'])
                cell.data_type = openpyxl.cell.cell.TYPE_STRING
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        else:
            writer.writerow([''])
            writer.writerow([f'Box {unpad_loc(box)}'] + [doc['title'] for doc in cfg.col_docs])
        for row in sorted(boxdict[box], key=lambda x: normalize_id(x[0])):
            if is_xlsx:
                nrow += 1
                for ncol, val in enumerate(row, start=1):
                    cell = ws.cell(row=nrow, column=ncol, value=val)
                    cell.data_type = openpyxl.cell.cell.TYPE_STRING
            else:
                writer.writerow(row)
    if is_xlsx:
        ws.column_dimensions["A"].width = 15  # Serial
        ws.column_dimensions["B"].width = 8  # Normal location
        ws.column_dimensions["C"].width = 50  # Title
        ws.column_dimensions["D"].width = 40  # Condition
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.oddHeader.center.text = _args.header
        ws.evenHeader.center.text = _args.header
        ws.oddFooter.left.text = "Page &P of &N"
        ws.evenFooter.left.text = "Page &P of &N"
        wb.save(_args.outfile)


def getparser() -> argparse.ArgumentParser:
    """
    Called either by getargs() in this file or by Sphinx.
    :return: an argparse.ArgumentParser object
    """
    parser = argparse.ArgumentParser(description='''
        Create a report of box contents with columns for the stocktake. The
        script uses a configuration file embedded in the Python code.''')
    parser.add_argument('infile', help=sphinxify('''
        Modes XML database file.
        ''', called_from_sphinx))
    parser.add_argument('outfile', help='''
        The output CSV or XLSX file.''')
    parser.add_argument('--header', default='', help='''
        Set the page heading.''')
    parser.add_argument('-s', '--short', action='store_true', help='''
        Only process one object. For debugging.''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary
        information.''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


called_from_sphinx = True


if __name__ == '__main__':
    assert sys.version_info >= (3, 11)
    called_from_sphinx = False
    if len(sys.argv) == 1:
        sys.argv.append('-h')
    _args = getargs(sys.argv)
    cfg_file = io.StringIO(CFG_STRING)
    cfg = Config(cfg_file, dump=_args.verbose > 1)
    is_xlsx = _args.outfile.lower().endswith('.xlsx')
    if is_xlsx:
        wb = Workbook()
        del wb[wb.sheetnames[0]]  # remove the default sheet
        ws = wb.create_sheet('Sheet1')
    else:
        encoding = 'utf-8-sig'
        csvfile = codecs.open(_args.outfile, 'w', encoding)
        outcsv = csv.writer(csvfile, delimiter=',')
        writer = csv.writer(outcsv)

    boxdict = defaultdict(list)
    main(cfg)
