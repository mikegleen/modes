"""
    Read the xlsx files in a folder and write the merged output to a single
    file, filtering on a column.
"""
import argparse
import os.path
import sys

import openpyxl.cell.cell
from openpyxl import Workbook

from utl.normalize import sphinxify


def onefile(filename, ws, prefix):
    global out_row_num
    infilepath = os.path.join(_args.indir, filename)
    wb1 = openpyxl.load_workbook(infilepath)
    ws1 = wb1.active
    in_row_num = 0
    for row in ws1.iter_rows():
        # Print the first heading row but skip the remaining ones.
        if not in_row_num:
            if not out_row_num:
                ws.append((cell.value for cell in row))
            in_row_num += 1
            continue
        out_row_num += 1
        row = [cell.value for cell in row]
        row[0] = f'{prefix}.{row[0]}'
        ws.append(row)
    return


def main():
    infiles = os.listdir(_args.indir)
    workbook = Workbook()
    ws = workbook.active
    for infile in infiles:
        prefix, suffix = os.path.splitext(infile)
        if suffix.lower() != '.xlsx':
            print(f'Skipping {infile}')
            continue
        onefile(infile, ws, prefix)
        workbook.save(_args.outfile)
        if _args.short:
            return


def getparser():
    parser = argparse.ArgumentParser(description=sphinxify('''

                ''', calledfromsphinx))
    parser.add_argument('indir', help='''
        Input folder with CSV or XLSX files.
        ''')
    parser.add_argument('outfile', help='''
        Output file to contain merged rows.
        ''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    parser.add_argument('-s', '--short', action='store_true', help='''
        Only process one object. For debugging.''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


calledfromsphinx = True


if __name__ == '__main__':
    calledfromsphinx = False
    assert sys.version_info >= (3, 9)
    out_row_num = 0
    if len(sys.argv) == 1:
        sys.argv.append('-h')
    _args = getargs(sys.argv)
    main()
