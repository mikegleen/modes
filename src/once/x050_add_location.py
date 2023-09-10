"""
    Create a new spreadsheet with the columns defined by NEWCOLS. If any of the
    columns exist in the old spreadsheet copy those.

    If the names don’t match exactly, map the new to the old name in
    NEW_TO_OLD_MAP.

    If the output file is .xlsx, all cells are set to number type "text".

"""
import argparse
import codecs
import csv
import sys

import openpyxl.cell.cell
from openpyxl import Workbook

from utl.normalize import sphinxify
from utl.readers import row_dict_reader
from utl.cfgutil import expand_idnum

NEWCOLS = ('Serial,Pages,Date,Person From,Person To,Org From,Org To,Type,'
           'Publ Name,Publ Date,Publ Page,Title,Author,Comment,'
           'Description,Location').split(',')

NEW_TO_OLD_MAP = {'Pages': 'Multiple Images',
                  'Person From': 'From',
                  'Person To': 'To'}

LOCS = """
G8 JB1204&6&11&13
G9 JB1203&8&9&7&28
G10 JB1202&18&19&20&21&10&12
G10 JB1205A
G11 JB1205B
"""
loclist = LOCS.split('\n')


def make_locdict():
    locdict = {}
    for row in loclist:
        if not row.strip():
            continue  # skip blank lines
        loc, assns = row.split()
        # print(assns)
        assnlist = expand_idnum(assns)
        for assn in assnlist:
            locdict[assn] = loc
        print(loc, assnlist)
    return locdict


def main():
    locdict = make_locdict()
    infilepath = _args.infile
    outfilepath = _args.outfile
    is_xlsx = outfilepath.lower().endswith('.xlsx')
    workbook = ws = outcsv = None  # Stop pycharm whining
    if is_xlsx:
        workbook = Workbook()
        ws = workbook.active
        for col_num, colname in enumerate(NEWCOLS, start=1):
            cell = ws.cell(row=1, column=col_num, value=colname)
            cell.data_type = openpyxl.cell.cell.TYPE_STRING
    else:
        encoding = 'utf-8-sig'
        csvfile = codecs.open(outfilepath, 'w', encoding)
        outcsv = csv.DictWriter(csvfile, fieldnames=NEWCOLS)
        outcsv.writeheader()
    row_num = 1
    for inrow in row_dict_reader(infilepath, _args.verbose):
        row_num += 1
        outrowdict = {}
        for col_num, colname in enumerate(NEWCOLS, start=1):
            if colname in inrow:
                value = inrow[colname]
            elif colname in NEW_TO_OLD_MAP:
                value = inrow[NEW_TO_OLD_MAP[colname]]
            else:
                value = ''
            if value is None:
                value = ''
            outrowdict[colname] = value
        if not ''.join(list(outrowdict.values())[1:]):  # empty row
            print(f'Skipping empty row {row_num}.')
            continue
        # print(outrowdict)
        assns = inrow['Serial'].split('.')  # JB001.1 -> ['JB001', '1']
        assn = assns[0]
        if assn in locdict:
            outrowdict['Location'] = locdict[assn]
        else:
            outrowdict['Location'] = 'N/A'
        if is_xlsx:
            for col_num, col_key in enumerate(NEWCOLS, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=outrowdict[col_key])
                cell.data_type = openpyxl.cell.cell.TYPE_STRING
                cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        else:
            outcsv.writerow(outrowdict)

    if is_xlsx:
        ws.freeze_panes = 'A2'
        workbook.save(outfilepath)
    return


def getparser():
    parser = argparse.ArgumentParser(description=sphinxify('''

                ''', calledfromsphinx))
    parser.add_argument('infile', help='''
        Input file with CSV or XLSX files.
        ''')
    parser.add_argument('outfile', help='''
        Output file.
        ''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


calledfromsphinx = True


if __name__ == '__main__':
    calledfromsphinx = False
    assert sys.version_info >= (3, 9)
    if len(sys.argv) == 1:
        sys.argv.append('-h')
    _args = getargs(sys.argv)
    main()
