"""

"""
import argparse
from openpyxl import load_workbook
import sys


def getparser():
    parser = argparse.ArgumentParser(description='''
        Remove line feeds from the cells in an Excel spreadsheet.''')
    parser.add_argument('infile', help='''
        The input XLSX file.''')
    parser.add_argument('outfile', help='''
        The output XLSX file.''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


if __name__ == '__main__':
    assert sys.version_info >= (3, 6)
    _args = getargs(sys.argv)
    wb = load_workbook(filename=_args.infile)
    ws = wb.active
    print(f'{ws.max_row} rows, {ws.max_column} columns')
    for nrow, row in enumerate(ws.iter_rows()):
        for ncol, col in enumerate(row):
            value = col.value
            if type(value) == str:
                col.value = value.replace('\n', ' ')
            if col.value != value:
                print(f'row: {nrow + 1}, col: {ncol + 1}', col.value)
    wb.save(_args.outfile)
