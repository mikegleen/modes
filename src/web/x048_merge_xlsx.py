"""
    Read the xlsx files in a folder and write the merged output to a single
    file, filtering on a column. Sort the output on a normalized accession number.
"""
import argparse
import os.path
import sys

import openpyxl.cell.cell
from openpyxl import Workbook

from utl.normalize import sphinxify, normalize_id


def onefile(filename, ws, prefix):
    global out_row_num, sheet
    # print(f'{filename=}')
    infilepath = os.path.join(_args.indir, filename)
    wb1 = openpyxl.load_workbook(infilepath)
    ws1 = wb1.active
    in_row_num = 0
    for row in ws1.iter_rows():
        # Print the first heading row but skip the remaining ones.
        if not in_row_num:
            if not out_row_num:
                ws.append((cell.value for cell in row))
            in_row_num += 1
            continue
        out_row_num += 1
        # row = [cell.value for cell in row]
        if not row[0].value:
            continue
        row[0].value = f'{prefix}.{row[0].value}'
        # print(f'{row[0].value=}')
        sheet.append((normalize_id(row[0].value), row))

    return


def put_ws(wb, ws):
    sheet.sort()
    for row_num, row in enumerate(sheet, start=2):
        for col_num, c in enumerate(row[1], start=1):
            cell = ws.cell(row=row_num, column=col_num, value=c.value)
            # cell.data_type = openpyxl.cell.cell.TYPE_STRING
            cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT

    # for row in sheet:
    #     out_row_num += 1
    #     ws.append(row[1])
    wb.save(_args.outfile)


def main():
    infiles = os.listdir(_args.indir)
    workbook = Workbook()
    worksheet = workbook.active
    for infile in infiles:
        prefix, suffix = os.path.splitext(infile)
        if prefix.startswith('~'):
            print('Warning: skipping open file:', infile)
            continue
        if suffix.lower() != '.xlsx' or prefix.lower() == 'template':
            print(f'Skipping {infile}')
            continue
        onefile(infile, worksheet, prefix)
        if _args.short:
            put_ws(workbook, worksheet)
            return
    put_ws(workbook, worksheet)


def getparser():
    parser = argparse.ArgumentParser(description=sphinxify('''

                ''', calledfromsphinx))
    parser.add_argument('indir', help='''
        Input folder with CSV or XLSX files.
        ''')
    parser.add_argument('outfile', help='''
        Output file to contain merged rows.
        ''')
    parser.add_argument('-v', '--verbose', type=int, default=1, help='''
        Set the verbosity. The default is 1 which prints summary information.
        ''')
    parser.add_argument('-s', '--short', action='store_true', help='''
        Only process one object. For debugging.''')
    return parser


def getargs(argv):
    parser = getparser()
    args = parser.parse_args(args=argv[1:])
    return args


calledfromsphinx = True


if __name__ == '__main__':
    calledfromsphinx = False
    assert sys.version_info >= (3, 9)
    out_row_num = 0
    sheet = []
    if len(sys.argv) == 1:
        sys.argv.append('-h')
    _args = getargs(sys.argv)
    main()
